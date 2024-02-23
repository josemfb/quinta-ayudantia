# pyinstaller -F -w -n "ayudantia" --add-data "logo.png:." --exclude-module "config.py" main.py

import datetime as dt
import locale
import os
import sys

import openpyxl
import openpyxl.styles
import pandas as pd

import config
import gui


def main() -> int:
    # options = ["Reportes mensuales (cuadro y parte)",
    #            "Reportes trimestrales (obligatorios y porcentajes)",
    #            "Configuración"]
    # while True:
    #     match gui.menu(options):
    #         case 1:
    #             return generate_monthly()
    #         case 2:
    #             return generate_quarterly()
    #         case 3:
    #             settings()
    #         case _:
    #             return 1
    return generate_monthly()


def generate_monthly() -> int:
    # Welcome
    gui.show_message(title="Reportes mensuales",
                     message="Se generarán de manera automática el cuadro mensual y el"
                             "\nparte de asistencia, a partir del archivo que se puede descargar"
                             "\ndesde Quintanet."
                             "\nA continuación deberá cargar el archivo .xls obtenido en"
                             "\nQuintanet. Es importante haber seleccionado las fechas"
                             "\ncorrespondientes a un solo mes, y haber desmarcado la opción"
                             "\n“Solo obligatorios”.",
                     button_text="Cargar archivo")

    # Ask user for Excel file path
    # Defaults to excel.html
    file = gui.get_file()
    if file == "":
        return 1

    # If file extension is .xls change it to .html
    if file[-4:] == ".xls":
        os.rename(file, file[:-4] + ".html")
        file = file[:-4] + ".html"

    # Load table to a pandas dataframe and drop the last seven columns
    # Those columns are statistics generated by Quintanet that we'll calculate again
    df = pd.DataFrame(pd.read_html(file)[0])
    df = df.drop(columns=["Oblig.", "Asist.", "Faltas", "Licencias", "% Oblig.", "% Sin Licencia",
                          "Series de 5 Faltas"])

    # Rename columns
    df = df.rename(columns={"Orden": "Ant.", "Voluntario.1": "Cargo"})

    # Sort by seniority and update that numbers
    df = df.sort_values("Ant.")
    df = df.reset_index(drop=True)
    for i in range(len(df.index)):
        df.at[i, "Ant."] = i + 1

    # Add lines for date, hours and type
    col_names = df.columns.tolist()

    # Review acts
    are_rog = False
    for act in col_names:
        day = f"{int(act[0:2]):02d}/{int(act[3:5]):02d}"
        time = f"{int(act[12:14]):02d}:{int(act[15:17]):02d}"

        # Detect ROG, change it to REG
        if act[19:] == "ROG":
            gui.show_message(title="Alerta: ROG",
                             message=f"Se detectó una ROG el {day} a las {time}."
                                     f"\nEl reglamento de Compañía señala que todas las reuniones"
                                     f"\ncitadas por el Directorio serán extraordinarias, por lo"
                                     f"\nque se cambió automáticamente de “Reunión Ordinaria"
                                     f"\nGeneral” a “Reunión a Extraordinaria”."
                             )
            # TODO - Update DF
            ...
            are_rog = True

        # Detect 10-1, ask for correct classification
        elif act[19:] == "10-1":
            options = ["10-1-1", "10-1-2", "10-1-3", "10-1-4",
                       "10-1-5", "10-1-6", "10-1-7", "10-1-8"]
            correct_act = gui.show_options(options, title="Alerta: 10-1",
                                           message=f"Se detectó un 10-1 el {day} a las {time}.\n"
                                                   f"Indique a que acto corresponde en realidad")
            # TODO - Update DF
            ...

        # Detect 10-6, ask for correct classification
        elif act[19:] == "10-6":
            options = ["10-6-1", "10-6-2", "10-6-3"]
            correct_act = gui.show_options(options, title="Alerta: 10-6",
                                           message=f"Se detectó un 10-6 el {day} a las {time}.\n"
                                                   f"Indique a que acto corresponde en realidad")
            # TODO - Update DF
            ...

        # Detect FDoM, ask if it's a FDoM or a FQ
        elif act[19:] == "FDoM":
            options = ["FDoM", "FQ"]
            correct_act = gui.show_options(options, title="Alerta: Funeral",
                                           message=f"Se detectó un funeral el {day} a las {time}.\n"
                                                   f"Indique a que acto corresponde en realidad")
            # TODO - Update DF
            ...

        # Detect 10-17-3, ask if it's a 10-0-6, 10-17 or 10-18
        elif act[19:] == "10-7-3":
            options = ["10-0-6", "10-17", "10-18"]
            correct_act = gui.show_options(options, title="Alerta: 10-7-3",
                                           message=f"Se detectó un 10-7-3 el {day} a las {time}.\n"
                                                   f"Indique a que acto corresponde en realidad")
            if correct_act == "10-17":
                options = ["10-17-0", "10-17-1", "10-17-2", "10-17-3", "10-17-4",
                           "10-17-5", "10-17-6", "10-17-7", "10-17-8"]
                correct_act = gui.show_options(options, title="Alerta: 10-7-3",
                                               message="Indique la subclasificación que corresponda")
            if correct_act == "10-18":
                options = ["10-18-0", "10-18-1", "10-18-2", "10-18-3", "10-18-4",
                           "10-18-5", "10-18-6", "10-18-7", "10-18-8"]
                correct_act = gui.show_options(options, title="Alerta: 10-7-3",
                                               message="Indique la subclasificación que corresponda")
            # TODO - Update DF
            ...

        # Detect DE, ask if it's a DE or EJ-G
        elif act[19:] == "DE":
            options = ["DE", "EJ-G"]
            correct_act = gui.show_options(options, title="Alerta: Delegación",
                                           message=f"Se detectó una delegación el {day} a las {time}.\n"
                                                   f"Indique a que acto corresponde en realidad")
            # TODO - Update DF
            ...
    col_names = df.columns.tolist()

    dates = ["", "", "Fecha"]
    hours = ["", "", "Hora"]
    acts = ["", "", "Acto"]
    bonus = ["", "", "ABH"]
    for col in col_names[3:]:
        dates.append(dt.date(int(col[6:10]), int(col[3:5]), int(col[0:2])))
        hours.append(dt.time(int(col[12:14]), int(col[15:17])).strftime("%H:%M"))
        acts.append(col[19:])
        bonus.append("")
    dates_dict = {col_names[i]: [dates[i]] for i in range(len(col_names))}
    dates_df = pd.DataFrame(dates_dict)
    hours_dict = {col_names[i]: [hours[i]] for i in range(len(col_names))}
    hours_df = pd.DataFrame(hours_dict)
    acts_dict = {col_names[i]: [acts[i]] for i in range(len(col_names))}
    acts_df = pd.DataFrame(acts_dict)
    bonus_dict = {col_names[i]: [bonus[i]] for i in range(len(col_names))}
    bonus_df = pd.DataFrame(bonus_dict)
    df = pd.concat([dates_df, hours_df, acts_df, bonus_df, df])

    # TODO - Ask about act length bonus
    ...

    seniority_20 = df[df["Voluntario"] == config.last_with_20]["Ant."].values[0]
    df["Ant."] = pd.to_numeric(df["Ant."])
    for series_name, _ in df.iloc[:, 3:].items():
        # Change “F” from volunteers with more than 20 years to “-”
        df[series_name] = df[series_name].mask(
            (df["Ant."] <= seniority_20) & (df[series_name] == "F"), "-", inplace=False
        )

        # Change “-” from honorary volunteers with less than 20 years to “F”
        if series_name[19:] in config.mandatory_for_honoraries:
            df[series_name] = df[series_name].mask(
                (df["Ant."] > seniority_20)
                & (df["Cargo"] == "Voluntario Activo")
                & (df[series_name] == "-"),
                "F", inplace=False
            )

        # Change “-” from officers to “F” (captain, lieutenants, assistants)
        if (series_name[19:] in config.dept_mandatory_acts
                or series_name[19:] in config.comp_mandatory_acts):
            df[series_name] = df[series_name].mask(
                ((df["Cargo"] == "Capitán") |
                 (df["Cargo"] == "Teniente Primero") |
                 (df["Cargo"] == "Teniente Segundo") |
                 (df["Cargo"] == "Teniente Tercero") |
                 (df["Cargo"] == "Ayudante")
                 )
                & (df[series_name] == "-"),
                "F", inplace=False
            )

        # Change “-” from officers to “F” (engineer)
        if ((series_name[19:] in config.dept_mandatory_acts
             or series_name[19:] in config.comp_mandatory_acts)
                and series_name[19:] not in ["FU", "RG"]):
            df[series_name] = df[series_name].mask(
                (df["Cargo"] == "Maquinista")
                & (df[series_name] == "-"),
                "F", inplace=False
            )

        # Change “-” from officers to “F” (secretary, treasurer, intendent)
        if ((series_name[19:] in config.dept_mandatory_acts
             or series_name[19:] in config.comp_mandatory_acts)
                and series_name[19:] not in ["FU", "RG", "EJ-GEN", "EJ", "10-34", "10-30"]):
            df[series_name] = df[series_name].mask(
                ((df["Cargo"] == "Secretario") |
                 (df["Cargo"] == "Tesorero") |
                 (df["Cargo"] == "Intendente")
                 )
                & (df[series_name] == "-"),
                "F", inplace=False
            )

    # Calculate statistics
    dept_mandatory = ["", "", "", "Ob. Gen."]
    comp_mandatory = ["", "", "", "Ob. Cía."]
    others = ["", "", "", "Ab."]
    total = ["", "", "", "Total"]

    abh = df.iloc[3, 3:].tolist()
    for n, item in enumerate(abh):
        if item == "":
            abh[n] = 1
        else:
            abh[n] = int(abh[n])

    sum_dept_mandatory = 0
    sum_comp_mandatory = 0
    sum_others = 0

    for _, row in df.iterrows():
        if row["Ant."] >= 1:
            series_count = 0
            for series_name, series in row.items():
                if series_name == "Ant." or series_name == "Voluntario" or series_name == "Cargo":
                    continue
                if series == "A":
                    if series_name[19:] in config.dept_mandatory_acts:
                        sum_dept_mandatory += 1
                        sum_others += abh[series_count] - 1
                    elif series_name[19:] in config.comp_mandatory_acts:
                        sum_comp_mandatory += 1
                        sum_others += abh[series_count] - 1
                    else:
                        sum_others += abh[series_count]
                series_count += 1
            dept_mandatory.append(sum_dept_mandatory)
            comp_mandatory.append(sum_comp_mandatory)
            others.append(sum_others)
            total.append(sum_dept_mandatory + sum_comp_mandatory + sum_others)
            sum_dept_mandatory = 0
            sum_comp_mandatory = 0
            sum_others = 0

    df["Ob. Gen."] = dept_mandatory
    df["Ob. Cía."] = comp_mandatory
    df["Ab."] = others
    df["Total"] = total

    df = df.copy()

    # Export excel: full monthly report
    name = f"{dates[3].year}-{dates[3].month:02d}"
    name = "Cuadro Mensual - " + name + ".xlsx"
    df.to_excel(name)

    # Open output to format
    wb = openpyxl.load_workbook(name)
    ws = wb.active

    # Remove extra columns
    ws.delete_cols(1)
    ws.delete_rows(1)

    # Define border to apply later
    border_color = openpyxl.styles.Side(border_style="thin", color='000000')
    border = openpyxl.styles.Border(left=border_color, right=border_color,
                                    top=border_color, bottom=border_color)

    # Add title
    ws.title = f"{dates[3].year}-{dates[3].month:02d}"
    ws["B1"] = "5.ª Compañía"
    ws["B2"] = "Cuadro Mensual"
    month = f"{dates[3].strftime("%B")} del {dates[3].year}"[0].upper()
    month += f"{dates[3].strftime("%B")} del {dates[3].year}"[1:]
    ws["B3"] = month
    ws["B1"].font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    ws["B2"].font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    ws["B3"].font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    ws["A1"].border = openpyxl.styles.Border(top=border_color)
    ws["B1"].border = openpyxl.styles.Border(top=border_color)
    ws["A1"].border = openpyxl.styles.Border(left=border_color)
    ws["A2"].border = openpyxl.styles.Border(left=border_color)
    ws["A3"].border = openpyxl.styles.Border(left=border_color)
    ws["A4"].border = openpyxl.styles.Border(left=border_color)
    ws["A1"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
    ws["A2"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
    ws["A3"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
    ws["A4"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
    ws["B1"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
    ws["B2"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
    ws["B3"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
    ws["B4"].fill = openpyxl.styles.PatternFill("solid", fgColor="004000")

    # Format dates, hours and acts
    ws["C1"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws["C2"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws["C3"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws["C4"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws["C1"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
    ws["C2"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
    ws["C3"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
    ws["C4"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
    ws["C1"].font = openpyxl.styles.Font(bold=True)
    ws["C2"].font = openpyxl.styles.Font(bold=True)
    ws["C3"].font = openpyxl.styles.Font(bold=True)
    ws["C4"].font = openpyxl.styles.Font(bold=True)
    ws["C1"].border = border
    ws["C2"].border = border
    ws["C3"].border = border
    ws["C4"].border = border

    col = "D"
    to_check = ws[f"{col}1"]
    while to_check.value is not None:
        ws[f"{col}1"].number_format = "dd/mm"
        ws[f"{col}2"].number_format = "hh:mm"
        ws[f"{col}4"].number_format = '"(+"####" ABH)"'
        ws[f"{col}1"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"{col}2"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"{col}3"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"{col}4"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws[f"{col}1"].font = openpyxl.styles.Font(bold=True)
        ws[f"{col}2"].font = openpyxl.styles.Font(bold=True)
        ws[f"{col}3"].font = openpyxl.styles.Font(bold=True)
        ws[f"{col}4"].font = openpyxl.styles.Font(bold=True, size=8)
        ws[f"{col}1"].border = border
        ws[f"{col}2"].border = border
        ws[f"{col}3"].border = border
        ws[f"{col}4"].border = border
        ws[f"{col}1"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
        ws[f"{col}2"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
        ws[f"{col}3"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
        ws[f"{col}4"].fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
        ws.column_dimensions[col].width = 7
        if col[-1:] == "Z":
            if len(col) == 1:
                col = "AA"
            else:
                col = f"{chr(ord(col[0]) + 1)}A"
        else:
            if len(col) == 1:
                col = f"{chr(ord(col[0]) + 1)}"
            else:
                col = f"{col[0]}{chr(ord(col[-1:]) + 1)}"
        to_check = ws[f"{col}1"]

    # Format statistics
    for i in range(4):
        ws.merge_cells(f"{col}1:{col}4")
        to_check = ws[f"{col}1"]
        to_check.fill = openpyxl.styles.PatternFill("solid", fgColor="004000")
        to_check.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        to_check.alignment = openpyxl.styles.Alignment(vertical="center", horizontal="center",
                                                       wrap_text=True)
        ws.column_dimensions[col].width = 7
        match i:
            case 0:
                ws[f"{col}1"] = "Ob. Gen."
            case 1:
                ws[f"{col}1"] = "Ob. Cía."
            case 2:
                ws[f"{col}1"] = "Ab."
            case 3:
                ws[f"{col}1"] = "Total"

        if col[-1:] == "Z":
            if len(col) == 1:
                col = "AA"
            else:
                col = f"{chr(ord(col[0]) + 1)}A"
        else:
            if len(col) == 1:
                col = f"{chr(ord(col[0]) + 1)}"
            else:
                col = f"{col[0]}{chr(ord(col[-1:]) + 1)}"

    # Format names and attendance
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 20

    row = 5
    while ws[f"A{row}"].value is not None:
        ws.row_dimensions[row].height = 30
        ws[f"B{row}"].font = openpyxl.styles.Font(bold=True)
        ws[f"A{row}"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws[f"B{row}"].alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
        ws[f"C{row}"].alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        ws[f"C{row}"].border = border

        col = "D"
        to_check = ws[f"{col}{row}"]
        while to_check.value is not None:
            to_check.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            to_check.border = border
            if to_check.value == "A":
                to_check.font = openpyxl.styles.Font(color="004000")
                to_check.fill = openpyxl.styles.PatternFill("solid", fgColor="C4D79B")
            elif to_check.value == "F":
                to_check.font = openpyxl.styles.Font(color="974706")
                to_check.fill = openpyxl.styles.PatternFill("solid", fgColor="FABF8F")
            elif to_check.value == "L":
                to_check.font = openpyxl.styles.Font(color="244062")
                to_check.fill = openpyxl.styles.PatternFill("solid", fgColor="95B3D7")
            elif to_check.value == "S" or to_check.value == "R":
                to_check.font = openpyxl.styles.Font(color="0D0D0D")
                to_check.fill = openpyxl.styles.PatternFill("solid", fgColor="A6A6A6")

            if col[-1:] == "Z":
                if len(col) == 1:
                    col = "AA"
                else:
                    col = f"{chr(ord(col[0]) + 1)}A"
            else:
                if len(col) == 1:
                    col = f"{chr(ord(col[0]) + 1)}"
                else:
                    col = f"{col[0]}{chr(ord(col[-1:]) + 1)}"
            to_check = ws[f"{col}{row}"]
        row += 1

    # Save formatted excel
    ws.freeze_panes = ws["D5"]
    wb.save(name)

    # TODO - Export excel: simple report
    ...
    # TODO - Format and color code simple report
    ...

    # Remove original file after finishing
    os.remove(file)

    return 0


def generate_quarterly():
    # TODO - Generate report with mandatory acts of the quarter
    ...


def settings():
    # TODO - Review and create new types of acts
    # TODO - Select which acts are mandatory
    # TODO - Select last volunteer with 20 years of service
    ...


if __name__ == "__main__":
    locale.setlocale(locale.LC_ALL, 'es_CL.utf8')
    sys.exit(main())
